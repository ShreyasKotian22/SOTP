from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
import psycopg2
from psycopg2.extras import RealDictCursor
import random
import smtplib
import hashlib
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime
import os
from sqlalchemy import text
from flask_mail import Mail, Message
import platform
from flask import request
import psutil, platform, time
from datetime import timedelta
from flask import g
from time import time
from collections import defaultdict

site_metrics = {
    "total_requests": 0,
    "errors": 0,
    "latencies": [],
    "endpoint_hits": defaultdict(int),
    "connected_users": set()
}

web_vital_metrics = defaultdict(list)



device_name = platform.node() or request.user_agent.platform
active_users = {}

def log_activity(email, action, details=None):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO activity_log (email, action, details)
            VALUES (%s, %s, %s)
        """, (email, action, details))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print("❌ Failed to log activity:", e)



def get_current_user_email():
    return session.get('user')



# Optional email config import
try:
    from email_config import EMAIL_HOST, EMAIL_PORT, EMAIL_USERNAME, EMAIL_PASSWORD
except ImportError:
    EMAIL_HOST = EMAIL_PORT = EMAIL_USERNAME = EMAIL_PASSWORD = None

app = Flask(__name__)
app.secret_key = 'your_secret_key'
JWT_SECRET_KEY = 'another_super_secret_key'


# === PostgreSQL Configs ===
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:shreyas@localhost/login_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

DB_CONFIG = {
    'dbname': 'login_db',
    'user': 'postgres',
    'password': 'shreyas',
    'host': 'localhost',
    'port': '5432'
}

def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)

# === Models ===
class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    real_role = db.Column(db.String)  # ✅ add this line
    role = db.Column(db.String)
    email = db.Column(db.String, unique=True)
    password = db.Column(db.String)
    registered_date = db.Column(db.DateTime)
    status = db.Column(db.String)


# === Utilities ===
verification_codes = {}

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def send_email(to_email, subject, body):
    try:
        message = f"Subject: {subject}\n\n{body}"
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
            server.starttls()
            server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
            server.sendmail(EMAIL_USERNAME, to_email, message)
    except Exception as e:
        print(f"Email error: {e}")
        raise

def get_orders():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, client_name, division, company_name, project_name, po_number, po_date, po_expiry_date,
               start_date, end_date, site, po_amount, status, issued_by,
               gst_percent, igst, cgst, sgst, tds_percent, tds_amount
        FROM orders
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{
        'order_id': row[0],
        'client_name': row[1],
        'division': row[2],
        'company': row[3],
        'project': row[4],
        'po_number': row[5],
        'po_date': row[6],
        'po_expiry_date': row[7],
        'start_date': row[8],
        'end_date': row[9],
        'site': row[10],
        'po_amount': row[11],
        'status': row[12],
        'issued_by': row[13],
        'gst_percent': row[14],
        'igst': row[15],
        'cgst': row[16],
        'sgst': row[17],
        'tds_percent': row[18],
        'tds_amount': row[19]
    } for row in rows]




# === Auth Routes ===
@app.route('/')
def index():
    return render_template('login.html')

    

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        password = request.form['password'].strip()
        hashed_password = hash_password(password)

        conn = get_db_connection()
        cur = conn.cursor()

        # Check if user exists and password matches
        cur.execute("""
            SELECT role, status FROM users
            WHERE LOWER(email) = %s AND password = %s
        """, (email, hashed_password))
        user = cur.fetchone()

        if user:
            role, status = user

            if status.lower() == 'enabled':
                session['user'] = email
                session['role'] = role

                # ✅ Reset kicked = FALSE on login
                cur.execute("""
                    UPDATE users
                    SET kicked = FALSE
                    WHERE LOWER(email) = %s
                """, (email,))

                # ✅ Log login activity
                cur.execute("""
                    INSERT INTO activity_log (email, action, details, timestamp)
                    VALUES (%s, %s, %s, NOW())
                """, (email, 'Login', 'Logged into the Portal'))

                conn.commit()
                cur.close()
                conn.close()

                return redirect(url_for('dashboard'))

            flash('Account is not enabled. Please contact administrator.', 'danger')
            cur.close()
            conn.close()
            return redirect(url_for('login'))

        flash('Invalid Credentials!', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('login'))

    return render_template('login.html')



@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        flash('Please log in first', 'danger')
        return redirect(url_for('index'))

    conn = get_db_connection()
    cur = conn.cursor()

    # Dashboard metrics
    cur.execute("SELECT COUNT(*) FROM clients")
    total_clients = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM orders")
    total_orders = cur.fetchone()[0]

    cur.execute("SELECT AVG(po_amount) FROM orders WHERE po_amount IS NOT NULL")
    avg_profit = cur.fetchone()[0] or 0

    # Use role column as display name
    cur.execute("SELECT role FROM users WHERE email = %s", (session['user'],))
    result = cur.fetchone()
    user_name = result[0] if result else "User"

    cur.close()
    conn.close()

    # Greeting
    hour = datetime.now().hour
    greeting = (
        "Good Morning," if 5 <= hour < 12 else
        "Good Afternoon," if 12 <= hour < 17 else
        "Good Evening," if 17 <= hour < 21 else
        "Hello,"
    )

    return render_template(
        'dashboard.html',
        email=session['user'],
        role=session['role'],
        user_name=user_name,
        greeting=greeting,
        total_clients=total_clients,
        total_orders=total_orders,
        average_profit=round(avg_profit, 2)
    )


@app.route('/logout')
def logout():
    user_email = session.get('user')
    user_role = session.get('role')  # Save before clearing session

    if user_email:
        log_activity(user_email, "Logout", "Logged Out of the Portal")

    session.clear()  # Clear session completely

    # Redirect based on the previous role
    if user_role == 'admin':
        return redirect('/adminlogin')
    else:
        return redirect('/login')





# === Forgot & Reset Password ===
@app.route('/forgot-password')
def forgot_password():
    return render_template('forgot.html')

@app.route('/send-code', methods=['POST'])
def send_code():
    data = request.get_json()
    email = data.get('email','').strip().lower()
    user = User.query.filter(func.lower(User.email)==email).first()
    if not user:
        return jsonify({'success': False, 'message': 'Email not registered.'})
    code = str(random.randint(100000, 999999))
    verification_codes[email] = code
    send_email(email, "Your Password Reset Code", f"Your code is: {code}")
    return jsonify({'success': True, 'message': 'Verification code sent.'})

@app.route('/verify-code', methods=['POST'])
def verify_code():
    data = request.get_json()
    email = data.get('email','').strip().lower()
    code = data.get('code','').strip()
    if verification_codes.get(email)!=code:
        return jsonify({'success': False, 'message': 'Invalid or missing code.'})
    verification_codes.pop(email, None)
    return jsonify({'success': True, 'message': 'Code verified.'})

@app.route('/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    email = data.get('email','').strip().lower()
    new_password = data.get('new_password','').strip()
    user = User.query.filter(func.lower(User.email)==email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found.'})
    user.password = hash_password(new_password)
    db.session.commit()
    verification_codes.pop(email, None)
    return jsonify({'success': True, 'message': 'Password reset successful.'})

# === Order Management ===
@app.route('/view-orders')
def view_orders():
    if 'user' not in session:
        flash('Please log in', 'danger')
        return redirect(url_for('index'))
    return render_template('view_sales_orders.html')

@app.route('/api/orders')
def api_orders():
    conn = get_db_connection()
    cur = conn.cursor()
    
    # --- THE FIX IS HERE: Added 'p_f' to the SELECT statement ---
    cur.execute("""
        SELECT id, client_name, division, company_name, project_name, po_number, po_date, po_expiry_date,
               start_date, end_date, site, po_amount, status, issued_by,
               gst_percent, igst, cgst, sgst, tds_percent, tds_amount,
               cess_percent, cess_amount, p_f
        FROM orders
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    orders = [{
        'order_id': row[0],
        'client_name': row[1],
        'division': row[2],
        'company': row[3],
        'project': row[4],
        'po_number': row[5],
        'po_date': row[6].isoformat() if row[6] else "",
        'po_expiry_date': row[7].isoformat() if row[7] else "",
        'start_date': row[8].isoformat() if row[8] else "",
        'end_date': row[9].isoformat() if row[9] else "",
        'site': row[10],
        'po_amount': row[11],
        'status': row[12],
        'issued_by': row[13],
        'gst_percent': row[14],
        'igst': row[15],
        'cgst': row[16],
        'sgst': row[17],
        'tds_percent': row[18],
        'tds_amount': row[19],
        'cess_percent': row[20],
        'cess_amount': row[21],
        # --- AND ADDED THE NEW 'p_f' KEY HERE ---
        'p_f': row[22] 
    } for row in rows]
    
    return jsonify(orders)

@app.route('/api/delete_order/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM orders WHERE id=%s", (order_id,))
        conn.commit()
        cur.close()
        conn.close()
        log_activity(session.get('user'), "Delete Order", f"Order ID: {order_id}")
        return jsonify({"message":"Order deleted"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

from flask import send_file
import tempfile
import openpyxl
from openpyxl.styles import Font

@app.route('/api/download/order/<int:order_id>')
def download_order_excel(order_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT client_name, division, company_name, project_name, sub_contractor_name,
               start_date, end_date, project_duration, site, po_number, po_amount,
               issued_by, po_date, po_expiry_date,
               gst_percent, gst_amount, cgst, sgst,
               tds_percent, tds_amount, status
        FROM orders
        WHERE id = %s
    """, (order_id,))
    
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return "Order not found", 404

    # Define headers in same order as the query
    columns = [
        "Client Name", "Division", "Company Name", "Project Name", "Sub Contractor Name",
        "Start Date", "End Date", "Project Duration", "Site", "PO Number", "PO Amount",
        "Issued By", "PO Date", "PO Expiry Date",
        "GST %", "GST Amount", "CGST", "SGST",
        "TDS %", "TDS Amount", "Status"
    ]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Order {order_id}"

    # Add headers
    for i, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = Font(bold=True)

    # Add data
    for i, value in enumerate(row, 1):
        ws.cell(row=2, column=i, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save to temp file and return
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)

    log_activity(session.get('user'), "Download Order", f"Order ID: {order_id}")

    return send_file(tmp.name, as_attachment=True, download_name=f'Order_{order_id}.xlsx')


@app.route('/add-order')
def add_order_form():
    return render_template('add.html')

@app.route('/submit', methods=['POST'])
def submit_order():
    data = request.get_json()
    print("🛠 Received order data:", data) 
    if not data:
        return jsonify({"error": "Invalid data"}), 400

    def clean(val):
        # Clean will now also handle numeric conversion for amounts
        if val in ("", None):
            return None
        return val

    # Required fields check remains the same
    required = ['clientName', 'division', 'companyName', 'projectName', 'poNumber', 'poAmount']
    for field in required:
        if not clean(data.get(field)):
            return jsonify({"error": f"Missing required field: {field}"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # --- THE FIX IS HERE ---
        # 1. Added 'p_f' to the list of columns
        # 2. Added one more '%s' placeholder for the p_f value
        cur.execute('''
            INSERT INTO orders (
                client_name, division, company_name, project_name, sub_contractor_name,
                start_date, end_date, project_duration, site, po_number, po_amount,
                issued_by, po_date, po_expiry_date,
                gst_percent, gst_amount, cgst, sgst, igst,
                tds_percent, tds_amount, cess_percent, cess_amount, status, p_f
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        ''', (
            clean(data.get('clientName')),
            clean(data.get('division')),
            clean(data.get('companyName')),
            clean(data.get('projectName')),
            clean(data.get('subContractorName')),
            clean(data.get('startDate')),
            clean(data.get('endDate')),
            clean(data.get('projectDuration')),
            clean(data.get('site')),
            clean(data.get('poNumber')),
            clean(data.get('poAmount')),
            clean(data.get('issuedBy')),
            clean(data.get('poDate')),
            clean(data.get('poExpiryDate')),
            clean(data.get('gstPercent')),
            clean(data.get('gstAmount')),
            clean(data.get('cgst')),
            clean(data.get('sgst')),
            clean(data.get('igst')),
            clean(data.get('tdsPercent')),
            clean(data.get('tdsAmount')),
            clean(data.get('cessPercent')),
            clean(data.get('cessAmount')),
            "Open",
            # --- THIS IS THE NEW VALUE ---
            clean(data.get('pfAmount')) # Get the pfAmount from the frontend data
        ))

        conn.commit()
        cur.close()
        conn.close()
        log_activity(session.get('user'), "Add Order", f"PO Number: {data.get('poNumber')}")
        return jsonify({"success": True}), 200
    except Exception as e:
        # It's helpful to print the error to your console for debugging
        print(f"❌ Error in /submit: {e}")
        return jsonify({"error": str(e)}), 500




    
# === Client Management ===
@app.route('/clients')
def clients():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, client_name, company_name, email, phone, billing_address, gst_no, pan
        FROM clients
    """)
    clients = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('addclients.html', clients=clients)

@app.route('/delete-client/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    try:
        conn = get_db_connection(); cur = conn.cursor()
        cur.execute("DELETE FROM clients WHERE id=%s", (client_id,))
        conn.commit(); cur.close(); conn.close()
        log_activity(session.get('user'), "Delete Client", f"Client ID: {client_id}")
        return jsonify({"success":True})
    except Exception as e:
        return jsonify({"success":False, "error":str(e)})

@app.route('/add-client', methods=['POST'])
def add_client():
    data = request.get_json()
    try:
        # Validate required fields
        required_fields = ['client_name', 'company_name', 'email', 'phone', 'billing_address', 'gst_no', 'pan']
        for field in required_fields:
            if not data.get(field):
                return jsonify({"error": f"Missing field: {field}"}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # Make sure the table exists
        cur.execute("""
            CREATE TABLE IF NOT EXISTS clients (
                id SERIAL PRIMARY KEY,
                client_name TEXT,
                company_name TEXT,
                email TEXT,
                phone TEXT,
                billing_address TEXT,
                gst_no TEXT,
                pan TEXT
            );
        """)

        # ✅ Check if email already exists
        cur.execute("SELECT id FROM clients WHERE email = %s", (data['email'],))
        existing_client = cur.fetchone()
        if existing_client:
            cur.close()
            conn.close()
            return jsonify({"duplicate": True, "message": "Client already exists with this email."}), 200


        # ✅ Insert new client
        cur.execute("""
            INSERT INTO clients (client_name, company_name, email, phone, billing_address, gst_no, pan)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['client_name'],
            data['company_name'],
            data['email'],
            data['phone'],
            data['billing_address'],
            data['gst_no'],
            data['pan']
        ))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        log_activity(session.get('user'), "Add Client", f"Email: {data['email']}")


        return jsonify({"message": "Client added successfully!", "client_id": new_id})
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/add-clients', methods=['GET'])
def add_client_form():
    return render_template('addclients.html')


    
@app.route('/api/update_order', methods=['PUT'])
def update_order():
    data = request.get_json()
    order_id = data.get('order_id')
    user_email = session.get('user')

    allowed_fields = [
        'client_name', 'company_name', 'division', 'project_name', 'sub_contractor_name',
        'start_date', 'end_date', 'project_duration', 'site',
        'po_number', 'po_amount', 'issued_by', 'po_date', 'po_expiry_date',
        'gst_percent', 'gst_amount', 'cgst', 'sgst',
        'tds_percent', 'tds_amount', 'status'
    ]

    if not order_id:
        return jsonify({'error': 'Order ID is required'}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # 1. Fetch current order data
        cur.execute(f"SELECT {', '.join(allowed_fields)} FROM orders WHERE id = %s", (order_id,))
        old_row = cur.fetchone()

        if not old_row:
            return jsonify({'error': 'Order not found'}), 404

        old_data = dict(zip(allowed_fields, old_row))

        # 2. Prepare update values
        updates = []
        values = []
        changes = []

        for field in allowed_fields:
            if field in data:
                new_val = data[field]
                old_val = old_data.get(field)
                if str(old_val).strip() != str(new_val).strip():
                    updates.append(f"{field} = %s")
                    values.append(new_val)
                    changes.append(f"{field}: '{old_val}' → '{new_val}'")

        if not updates:
            return jsonify({'message': 'No changes detected'}), 200

        values.append(order_id)

        # 3. Perform update
        cur.execute(f"""
            UPDATE orders
            SET {', '.join(updates)}
            WHERE id = %s
        """, tuple(values))
        conn.commit()

        # 4. Log activity
        change_str = "; ".join(changes)
        log_activity(user_email, "Edit Order", f"Order ID: {order_id}", changes=change_str)

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()

@app.route('/api/auto_update_status', methods=['POST'])
def auto_update_status():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        today = datetime.today().date()

        # Update status to 'Delayed' where po_expiry_date has passed and status is not already 'Delayed'
        cur.execute("""
            UPDATE orders
            SET status = 'Delayed'
            WHERE po_en_date < %s AND status != 'Delayed'
        """, (today,))
        updated_rows = cur.rowcount

        conn.commit()
        return jsonify({'success': True, 'message': f'{updated_rows} orders updated to Delayed'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()


@app.route("/update-client", methods=["PUT"])
def update_client():
    data = request.get_json()
    client_id = data.get("client_id")
    email = session.get('user')

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # 1. Fetch existing client data
        cur.execute("SELECT company_name, client_name, email, phone, gst_no, billing_address, pan FROM clients WHERE id = %s", (client_id,))
        old_data = cur.fetchone()

        if not old_data:
            return jsonify({"success": False, "error": "Client not found"}), 404

        # Map old data to dictionary
        old_fields = [
            "company_name", "client_name", "email", "phone",
            "gst_no", "billing_address", "pan"
        ]
        old_data_dict = dict(zip(old_fields, old_data))

        # 2. Compare old and new, collect changes
        changes = []
        for field in old_fields:
            old_val = old_data_dict[field]
            new_val = data.get(field)
            if str(old_val).strip() != str(new_val).strip():
                changes.append(f"{field}: '{old_val}' → '{new_val}'")

        # 3. Perform update
        cur.execute("""
            UPDATE clients
            SET company_name = %s,
                client_name = %s,
                email = %s,
                phone = %s,
                gst_no = %s,
                billing_address = %s,
                pan = %s
            WHERE id = %s
        """, (
            data["company_name"],
            data["client_name"],
            data["email"],
            data["phone"],
            data["gst_no"],
            data["billing_address"],
            data["pan"],
            client_id
        ))
        conn.commit()

        # 4. Log the update with changes
        change_str = "; ".join(changes) if changes else "No changes"
        log_activity(email, "Edit Client", f"Client ID: {client_id}", changes=change_str)

        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)})
    finally:
        cur.close()
        conn.close()


import pandas as pd


import psycopg2
import io
from flask import send_file

@app.route("/download-client/<path:email>")
def download_client_excel(email):
    conn = psycopg2.connect(
        host="localhost",
        database="login_db",
        user="postgres",
        password="shreyas"
    )
    cur = conn.cursor()

    # Remove remarks from SELECT
    cur.execute("""
        SELECT client_name, company_name, email, phone, billing_address, gst_no, id, pan
        FROM clients
        WHERE email = %s
    """, (email,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return "Client not found", 404

    columns = [
        "Client Name", "Company Name", "Email", "Phone", "Billing Address",
        "GST No", "Client ID", "PAN"
    ]

    import pandas as pd
    import io
    from flask import send_file

    df = pd.DataFrame([row], columns=columns)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Client Details')

    output.seek(0)
    safe_email = email.replace("@", "_at_").replace(".", "_")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Client_{safe_email}_Details.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route("/download-client/<int:client_id>")
def download_client(client_id):
    import pandas as pd
    import io
    from flask import send_file

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT company_name, client_name, email, phone, billing_address, gst_no, pan, id
        FROM clients WHERE id = %s
    """, (client_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return "Client not found", 404

    columns = [
        "Company Name", "Client Name", "Email", "Phone", "Billing Address",
        "GST No", "PAN", "Client ID"
    ]
    df = pd.DataFrame([row], columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Client Details')
    output.seek(0)
    log_activity(session.get('user'), "Download Client", f"Client ID: {client_id}")
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Client_{client_id}_Details.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route("/download-all-orders")
def download_all_orders():
    conn = psycopg2.connect(
        host="localhost",
        database="login_db",
        user="postgres",
        password="shreyas"
    )
    cur = conn.cursor()

    cur.execute("""
        SELECT client_name, division, company_name, project_name, sub_contractor_name,
               start_date, end_date, project_duration, site, po_number, po_amount,
               issued_by, po_date, po_expiry_date,
               gst_percent, gst_amount, cgst, sgst,
               tds_percent, tds_amount, status
        FROM orders
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        return "No orders found", 404

    # Define headers
    columns = [
        "Client Name", "Division", "Company Name", "Project Name", "Sub Contractor Name",
        "Start Date", "End Date", "Project Duration", "Site", "PO Number", "PO Amount",
        "Issued By", "PO Date", "PO Expiry Date",
        "GST %", "GST Amount", "CGST", "SGST",
        "TDS %", "TDS Amount", "Status"
    ]

    # Create DataFrame and write to Excel
    df = pd.DataFrame(rows, columns=columns)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="All Orders")

    output.seek(0)
    log_activity(session.get('user'), "Download All Orders")

    return send_file(
        output,
        as_attachment=True,
        download_name="All_Orders.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/api/clients')
def api_clients():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT client_name, company_name, email, phone, gst_no, pan, billing_address
        FROM clients
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    clients = []

    for row in rows:
        clients.append({
            "clientName": row[0],
            "companyName": row[1],
            "email": row[2],
            "phone": row[3],
            "gst": row[4],
            "pan": row[5].strip() if row[5] else "",
            "address": row[6]
        })

    return jsonify(clients)

@app.route("/get-clients")
def get_clients():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT client_name, company_name, email, phone, gst_no, pan, billing_address
            FROM clients
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        clients = {}
        for row in rows:
            name = row[0]
            clients[name] = {
                "clientName": row[0],
                "companyName": row[1],
                "email": row[2],
                "phone": row[3],
                "gst": row[4],
                "pan": row[5],
                "address": row[6]
            }

        return jsonify(clients)
    except Exception as e:
        print("❌ Error in /get-clients:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/report')
def report():
    if 'user' not in session:
        return redirect('/login')
    return render_template('reports.html')




#//////////////ROHAN'S SAMRAJYA///////////////////////////////

@app.route('/api/dashboard-metrics')
def get_dashboard_metrics():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')

        conn = get_db_connection()
        cur = conn.cursor()

        # Build date filter for orders table
        order_filter = ""
        order_params = []

        if from_date and to_date:
            order_filter = "WHERE po_date BETWEEN %s AND %s"
            order_params = [from_date, to_date]

        # Total jobs
        cur.execute(f"SELECT COUNT(*) FROM orders {order_filter}", order_params)
        total_jobs = cur.fetchone()[0]

        # Total sales value
        cur.execute(f"SELECT SUM(po_amount) FROM orders {order_filter}", order_params)
        total_sales_value = cur.fetchone()[0] or 0

        cur.execute(f"SELECT COUNT(DISTINCT company_name) FROM orders {order_filter}", order_params)
        total_companies = cur.fetchone()[0]

        cur.execute(f"SELECT COUNT(DISTINCT client_name) FROM clients {order_filter}", order_params)
        total_clients = cur.fetchone()[0]


        cur.close()
        conn.close()

        return jsonify({
            "total_jobs": total_jobs,
            "total_companies": total_companies,
            "total_clients": total_clients,
            "total_sales_value": float(total_sales_value)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500




@app.route('/api/job-status-metrics')
def get_job_status_metrics():
    try:
        from_date = request.args.get('from')
        to_date = request.args.get('to')

        conn = get_db_connection()
        cur = conn.cursor()

        date_filter = ""
        params = []

        if from_date and to_date:
            date_filter = "AND po_date BETWEEN %s AND %s"
            params = [from_date, to_date]

        statuses = ["open", "in progress", "on hold", "delayed", "closed", "cancelled"]
        status_counts = {}

        # Get counts and values for each status
        for status in statuses:
            # Count
            cur.execute(f"""
                SELECT COUNT(*) FROM orders
                WHERE LOWER(status) = %s {date_filter}
            """, [status] + params)
            count = cur.fetchone()[0]
            status_key = status.replace(" ", "_")
            status_counts[status_key] = count

            # Value
            cur.execute(f"""
                SELECT COALESCE(SUM(po_amount), 0) FROM orders
                WHERE LOWER(status) = %s {date_filter}
            """, [status] + params)
            value = cur.fetchone()[0]
            status_counts[f"{status_key}_value"] = float(value)

        cur.close()
        conn.close()

        return jsonify(status_counts)
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/api/jobs-by-status')
def get_jobs_by_status():
    from flask import request
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    status = request.args.get('status')

    conn = get_db_connection()
    cur = conn.cursor()

    query = "SELECT id, client_name, company_name, project_name, po_amount, status, issued_by FROM orders WHERE 1=1"
    params = []

    if status:
        query += " AND LOWER(status) = %s"
        params.append(status.lower())

    if from_date and to_date:
        query += " AND po_date BETWEEN %s AND %s"
        params.extend([from_date, to_date])

    cur.execute(query, params)
    rows = cur.fetchall()

    results = []
    for row in rows:
        results.append({
            "id": row[0],
            "client": row[1],
            "company": row[2],
            "project": row[3],
            "po_amount": float(row[4]),
            "status": row[5],
            "issued_by": row[6]
        })

    cur.close()
    conn.close()

    return jsonify(results)

@app.route("/api/order-details", methods=["POST"])
def get_order_details():
    from datetime import datetime, timedelta
    data = request.get_json()
    group_by = data.get("group_by")
    value = data.get("value")
    interval = data.get("interval", "Monthly")

    if not group_by or value is None:
        return jsonify([])

    conn = get_db_connection()
    cur = conn.cursor()

    if group_by == "po_date":
        try:
            value_date = datetime.strptime(value, "%Y-%m-%d")
        except Exception:
            cur.close()
            conn.close()
            return jsonify([])

        from dateutil.relativedelta import relativedelta

        if interval == "Monthly":
            start = value_date.replace(day=1)
            end = start + relativedelta(months=1)
                

        elif interval == "Quarterly":
            quarter = (value_date.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            start = value_date.replace(month=start_month, day=1)
            if start_month + 3 > 12:
                end = start.replace(year=start.year + 1, month=1, day=1)
            else:
                end = start.replace(month=start_month + 3, day=1)

        elif interval == "Yearly":
            start = value_date.replace(month=1, day=1)
            end = start.replace(year=start.year + 1)

        elif interval == "Weekly":
            start = value_date  # ISO week Monday already
            end = start + timedelta(days=7)

        else:
            cur.close()
            conn.close()
            return jsonify([])

        query = """
            SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                   start_date, end_date, site, po_number, po_amount, issued_by,
                   po_date, status, project_duration, po_expiry_date
            FROM orders
            WHERE DATE(po_date) >= %s AND DATE(po_date) < %s
        """
        cur.execute(query, (start.date(), end.date()))

    else:
        # For entity groupings (e.g., company_name, client_name, etc.)
        query = f"""
            SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                   start_date, end_date, site, po_number, po_amount, issued_by,
                   po_date, status, project_duration, po_expiry_date
            FROM orders
            WHERE {group_by} = %s
        """
        cur.execute(query, (value,))

    rows = cur.fetchall()
    cols = [desc[0] for desc in cur.description]
    result = [dict(zip(cols, row)) for row in rows]

    cur.close()
    conn.close()
    return jsonify(result)




#///////////////END OF ROHAN'S SAMRAJYA////////////////////////
#**************************************************************
#////////////////////////////////////////////////////////////////

@app.route('/api/performance-data', methods=['POST'])
def poperformance_data():
    
    from collections import defaultdict
    from datetime import datetime, date

    filters = request.get_json()

    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])
    interval = filters.get("interval", "Monthly")
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    group_by = filters.get("group_by", "period")

    def is_valid_date(date_str):
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except:
            return False

    valid_from = is_valid_date(from_date)
    valid_to = is_valid_date(to_date)

    where_clauses = []
    params = []

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            # ✅ Determine default from/to dates
            if not valid_from or not valid_to:
                cur.execute("SELECT MIN(po_date) FROM orders")
                min_po_date = cur.fetchone()[0]
                default_from = min_po_date if min_po_date else date.today()
                default_to = date.today()

                if not valid_from:
                    from_date = default_from.strftime("%Y-%m-%d")
                if not valid_to:
                    to_date = default_to.strftime("%Y-%m-%d")
            else:
                default_from = datetime.strptime(from_date, "%Y-%m-%d").date()
                default_to = datetime.strptime(to_date, "%Y-%m-%d").date()

            # Apply filters
            if company:
                where_clauses.append("company_name = ANY(%s)")
                params.append(company)
            if client:
                where_clauses.append("client_name = ANY(%s)")
                params.append(client)
            if division:
                where_clauses.append("division = ANY(%s)")
                params.append(division)
            if subcontractor:
                where_clauses.append("sub_contractor_name = ANY(%s)")
                params.append(subcontractor)
            if site:
                where_clauses.append("site = ANY(%s)")
                params.append(site)

            where_clauses.append("po_date BETWEEN %s AND %s")
            params.extend([from_date, to_date])

            query = """
                SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                       start_date, end_date, site, po_number, po_amount, issued_by,
                       po_date, status, project_duration, po_expiry_date
                FROM orders
                WHERE TRUE
            """
            if where_clauses:
                query += " AND " + " AND ".join(where_clauses)

            cur.execute(query, params)
            rows = cur.fetchall()
            cols = [desc[0] for desc in cur.description]
            results = []
            for row in rows:
                record = dict(zip(cols, row))
                record["order_id"] = record.pop("id", None)
                results.append(record)

    entity_groupings = [
        "company_name", "client_name", "division", "site", "sub_contractor_name"
    ]

    if group_by in entity_groupings:
        # For entity-based grouping
        output = []
        for row in results:
            full_entity = row.get(group_by) or "Unknown"
            short_entity = " ".join([word[0].upper() for word in full_entity.split() if word and word[0].isalpha()])
            date_obj = row.get("po_date")

            if date_obj:
                if interval == "Weekly":
                    time_label = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
                elif interval == "Quarterly":
                    time_label = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
                elif interval == "Yearly":
                    time_label = f"{date_obj.year}"
                else:
                    time_label = date_obj.strftime("%b %Y")
                combined_label = f"{short_entity} - {time_label}"
            else:
                combined_label = short_entity

            output.append({
                "period": combined_label,
                "po": row["po_number"],
                "y": float(row["po_amount"] or 0),
                "project": row.get("project_name") or "-",
                "entity_name": full_entity,
                "status": row.get("status", "Unknown"),
                "po_date": date_obj.strftime("%Y-%m-%d") if date_obj else None
            })

        def extract_date(label):
            try:
                timeline = label.split(" - ")[-1]
                if interval == "Weekly":
                    year, week = timeline.split(" W")
                    return datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w")
                elif interval == "Quarterly":
                    q, y = timeline.split()
                    month = (int(q[1]) - 1) * 3 + 1
                    return datetime(int(y), month, 1)
                elif interval == "Yearly":
                    return datetime(int(timeline), 1, 1)
                else:
                    return datetime.strptime(timeline, "%b %Y")
            except:
                return datetime.max

        output.sort(key=lambda x: extract_date(x["period"]))

        return jsonify({
            "data": output,
            "records": results,
            "default_from_date": default_from.strftime("%Y-%m-%d"),
            "default_to_date": default_to.strftime("%Y-%m-%d")
        })

    # Timeline grouping with status
    summary = defaultdict(lambda: {"total_sales": 0, "sales_count": 0, "status_counts": defaultdict(int)})
    for row in results:
        date_obj = row.get("po_date")
        if not date_obj:
            continue
        if interval == "Weekly":
            key = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
        elif interval == "Quarterly":
            key = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
        elif interval == "Yearly":
            key = f"{date_obj.year}"
        else:
            key = date_obj.strftime("%b %Y")

        status = (row.get("status") or "Unknown").lower()
        summary[key]["total_sales"] += float(row["po_amount"] or 0)
        summary[key]["sales_count"] += 1
        summary[key]["status_counts"][status] += 1

    output = []
    for period, data in sorted(summary.items()):
        output.append({
            "period": period,
            "total_sales": data["total_sales"],
            "sales_count": data["sales_count"],
            "statuses": dict(data["status_counts"])
        })

    return jsonify({
        "data": output,
        "records": results,
        "default_from_date": default_from.strftime("%Y-%m-%d"),
        "default_to_date": default_to.strftime("%Y-%m-%d")
    })




@app.route('/api/filter-options')
def filter_options():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    # Build WHERE clause for timeline
    where_clauses = []
    params = []

    if from_date and to_date:
        where_clauses.append("po_date BETWEEN %s AND %s")
        params.extend([from_date, to_date])

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT DISTINCT company_name FROM orders {where_sql} ORDER BY company_name ASC;", params)
            company_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT client_name FROM orders {where_sql} ORDER BY client_name ASC;", params)
            client_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT sub_contractor_name FROM orders WHERE sub_contractor_name IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY sub_contractor_name ASC",
                        params if from_date and to_date else [])
            subcontractors = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT site FROM orders WHERE site IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY site ASC",
                        params if from_date and to_date else [])
            sites = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT division FROM orders {where_sql} ORDER BY division ASC;", params)
            divisions = [row[0] for row in cur.fetchall()]

    return jsonify({
        "company_names": company_names,
        "client_names": client_names,
        "subcontractors": subcontractors,
        "sites": sites,
        "divisions": divisions
    })


@app.route("/api/filter-options-dependent", methods=["POST"])
def filter_options_dependent():
    filters = request.json or {}
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])

    where = []
    params = []

    if company:
        where.append("company_name = ANY(%s)")
        params.append(company)
    if client:
        where.append("client_name = ANY(%s)")
        params.append(client)
    if division:
        where.append("division = ANY(%s)")
        params.append(division)
    if subcontractor:
        where.append("sub_contractor_name = ANY(%s)")
        params.append(subcontractor)
    if site:
        where.append("site = ANY(%s)")
        params.append(site)
    if from_date:
        where.append("po_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("po_date <= %s")
        params.append(to_date)

    query = """
        SELECT DISTINCT client_name, division, sub_contractor_name, site, company_name
        FROM orders
    """
    if where:
        query += " WHERE " + " AND ".join(where)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Extract unique values for each field
    clients = sorted(set(r[0] for r in rows if r[0]))
    divisions = sorted(set(r[1] for r in rows if r[1]))
    subcontractors = sorted(set(r[2] for r in rows if r[2]))
    sites = sorted(set(r[3] for r in rows if r[3]))
    companies = sorted(set(r[4] for r in rows if r[4]))

    return jsonify({
        "client_names": clients,
        "divisions": divisions,
        "subcontractors": subcontractors,
        "sites": sites,
        "company_names": companies
    })



@app.route("/api/po-date-range")
def get_po_date_range():
    from datetime import date
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT MIN(po_date) FROM orders")
        min_date = cur.fetchone()[0]
        cur.close()
        conn.close()

        return jsonify({
            "min_date": min_date.strftime('%Y-%m-%d') if min_date else date.today().strftime('%Y-%m-%d'),
            "max_date": date.today().strftime('%Y-%m-%d')  # ✅ always today
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500




import hashlib
from flask import jsonify

@app.route("/api/po-colors")
def get_po_colors():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT po_number FROM orders")  # adjust if needed
    po_numbers = [row[0] for row in cur.fetchall() if row[0]]

    cur.close()
    conn.close()

    def hash_color(po):
        # Use a stable hash for consistent color, limit to 6 hex chars
        return "#" + hashlib.md5(po.encode()).hexdigest()[:6]

    po_colors = {po: hash_color(po) for po in po_numbers}

    return jsonify(po_colors)



from datetime import date

@app.route('/api/expired-pos-count')
def expired_pos_count():
    from flask import request
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    conn = get_db_connection()
    cur = conn.cursor()

    today = date.today()
    query = "SELECT COUNT(*), COALESCE(SUM(po_amount), 0) FROM orders WHERE po_expiry_date < %s"
    params = [today]

    # Optional date filter
    if from_date and to_date:
        query += " AND po_date BETWEEN %s AND %s"
        params.extend([from_date, to_date])

    cur.execute(query, params)
    count, value = cur.fetchone()
    cur.close()
    conn.close()

    return jsonify({
        "expired_pos_count": count,
        "expired_pos_value": float(value)
    })

import pandas as pd
import io
from flask import send_file

@app.route('/api/orders/export')
def export_orders():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM orders")
    rows = cur.fetchall()
    colnames = [desc[0] for desc in cur.description]
    cur.close()
    conn.close()

    df = pd.DataFrame(rows, columns=colnames)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Orders')
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="all_sales_orders.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

from flask import request, send_file
import pandas as pd
import io

@app.route('/api/orders/export_filtered', methods=['POST'])
def export_filtered_orders():
    data = request.get_json()
    order_ids = data.get('order_ids', [])
    if not order_ids:
        df = pd.DataFrame()
    else:
        conn = get_db_connection()
        cur = conn.cursor()
        # Convert all IDs to int (if your id column is integer)
        order_ids = [int(x) for x in order_ids]
        sql = "SELECT * FROM orders WHERE id = ANY(%s)"  # <-- FIXED HERE
        cur.execute(sql, (order_ids,))
        rows = cur.fetchall()
        colnames = [desc[0] for desc in cur.description]
        cur.close()
        conn.close()
        df = pd.DataFrame(rows, columns=colnames) if rows else pd.DataFrame(columns=colnames)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Orders')
    output.seek(0)
    log_activity(session.get('user'), "Download Selected Orders", f"Order IDs: {order_ids}")

    return send_file(
        output,
        as_attachment=True,
        download_name="sales_orders.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/api/company-names')
def get_company_names():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT company_name
        FROM clients
        WHERE company_name IS NOT NULL
        ORDER BY company_name ASC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    company_names = [row[0] for row in rows]
    return jsonify(company_names)


def has_permission_to_edit(current_role, target_role):
    """
    Checks if a user has permission to edit/delete another user.
    - Superadmin can manage Admins and Users.
    - Admin can manage Users.
    - No one can manage a Superadmin.
    """
    if current_role == "Superadmin":
        return target_role in ["Admin", "User"]
    elif current_role == "Admin":
        return target_role == "User"
    return False


#////////////////////////////////////////////////////////////////
from flask import Flask, render_template
from flask_apscheduler import APScheduler
from datetime import datetime, timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import psycopg2



# 🔶 2. Scheduler config
class Config:
    SCHEDULER_API_ENABLED = True

# 🔷 3. Attach config and start scheduler
app.config.from_object(Config())
scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()


# 🔔 Daily Scheduler at 11:09 AM
@scheduler.task('cron', hour=11, minute=27)
def notify_po_expiry():
    with app.app_context():  # Required for render_template
        conn = psycopg2.connect("dbname=login_db user=postgres password=shreyas host=localhost")
        cur = conn.cursor()

        target_date = (datetime.now() + timedelta(days=7)).date()
        cur.execute("""
            SELECT po_number, po_expiry_date, client_name, project_name
            FROM orders
            WHERE po_expiry_date = %s
        """, (target_date,))
        expiring_pos = cur.fetchall()

        if expiring_pos:
            cur.execute("SELECT email FROM users")
            users = [row[0] for row in cur.fetchall()]

            for po in expiring_pos:
                po_number, expiry_date, client_name, project = po

                html = render_template(
                    'po_expiry_email.html',
                    po_number=po_number,
                    client_name=client_name,
                    project_name=project,
                    po_expiry_date=expiry_date.strftime("%d %b %Y"),
                    today=datetime.now().strftime("%d %b %Y")
                )

                for email in users:
                    send_email(email, " PO Expiry Notice", html)

        cur.close()
        conn.close()
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import smtplib

def send_email(to_email, subject, html_content):
    msg = MIMEMultipart('related')
    msg['Subject'] = Header(subject, 'utf-8')  # Supports emojis
    msg['From'] = 'shreyas22032004@gmail.com'
    msg['To'] = to_email

    # HTML body
    msg_alternative = MIMEMultipart('alternative')
    msg.attach(msg_alternative)
    msg_alternative.attach(MIMEText(html_content, 'html', 'utf-8'))

    # Inline logo
    with open('static/assets/LOGO.png', 'rb') as f:
        img = MIMEImage(f.read())
        img.add_header('Content-ID', '<logo>')
        img.add_header('Content-Disposition', 'inline', filename='logo.png')
        msg.attach(img)

    # ✅ FIXED: use send_message, not sendmail
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('shreyas22032004@gmail.com', 'wvcnpazhztcutrmb')  # app password
        server.send_message(msg)






app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'shreyas22032004@gmail.com'        # Replace
app.config['MAIL_PASSWORD'] = 'wvcnpazhztcutrmb'      # Replace with App Password
app.config['MAIL_DEFAULT_SENDER'] = 'shreyas22032004@gmail.com'  # Optional fallback

mail = Mail(app)



@app.route("/share-order-pdf", methods=["POST"])
def share_order_pdf():
    recipient = request.form.get("recipient")
    pdf_file = request.files.get("pdf")

    current_user_email = get_current_user_email()
    if not current_user_email:
        return jsonify({"success": False, "error": "User not logged in"}), 401

    if not recipient or not pdf_file:
        return jsonify({"success": False, "error": "Missing data"}), 400

    msg = Message(
        subject="Order Summary PDF",
        sender=current_user_email,
        recipients=[recipient],
        body="Please find the attached Order Summary PDF."
    )
    msg.attach("order-summary.pdf", "application/pdf", pdf_file.read())
    mail.send(msg)

    return jsonify({"success": True})

@app.route('/send-pdf-email', methods=['POST'])
def send_pdf_email():
    to_email = request.form['to']
    pdf_file = request.files['pdf']

    msg = Message(subject="Your Order Summary",
                  sender="noreply@yourdomain.com",
                  recipients=[to_email])
    msg.body = "Please find attached the order summary."

    msg.attach(pdf_file.filename, "application/pdf", pdf_file.read())

    try:
        mail.send(msg)
        return jsonify(success=True)
    except Exception as e:
        print("Email send error:", e)
        return jsonify(success=False), 500



#############################--Deon's Samrajya--####################################
@app.route('/get_sub_contractors')
def get_sub_contractors():
    # Fetch sub contractor company names from your database
    sub_contractors = [ ... ]  # Replace with actual DB query
    return jsonify(sub_contractors)


  


@app.route('/adminlogin')
def admin_login():
    return render_template('adminlogin.html')

from flask import request, jsonify, session
import random, smtplib
from datetime import datetime, timedelta

from flask import request, jsonify, session
import random, smtplib
from datetime import datetime, timedelta

@app.route('/send-admin-otp', methods=['POST'])
def send_admin_otp():
    data = request.get_json()
    email = data.get('email')

    if not email:
        return jsonify(success=False, message="Email is required")

    # --- THE FIX IS HERE ---
    # The query now checks if the user's role is IN the list ['Admin', 'Superadmin']
    user = db.session.execute(
        db.select(User).where(
            func.lower(User.email) == email.lower(),
            User.real_role.in_(['Admin', 'Superadmin']), # CORRECTED LINE
            func.lower(User.status) == 'enabled'
        )
    ).scalar()

    if not user:
        # This message is now accurate, as it fails for non-admins or disabled accounts
        return jsonify(success=False, message="Email not found or is not an authorized administrator.")

    # Generate and send OTP as before
    otp = str(random.randint(100000, 999999))
    session['admin_otp'] = otp
    session['admin_email'] = email
    # Set a 2-minute expiry for the OTP for better security
    session['otp_expiry'] = (datetime.utcnow() + timedelta(minutes=2)).isoformat()

    try:
        send_email(email, "Your Admin Panel OTP", f"Your one-time password is: {otp}")
        log_activity(email, "Admin OTP Request", "Requested an OTP for Admin Panel login.")
        return jsonify(success=True)
    except Exception as e:
        print(f"❌ Failed to send OTP email to {email}: {e}")
        return jsonify(success=False, message="Could not send OTP email. Please try again later.")


@app.route('/verify-admin-otp', methods=['POST'])
def verify_admin_otp():
    data = request.get_json()
    email = data.get('email')
    otp = data.get('otp')

    if email != session.get('admin_email'):
        return jsonify(success=False, message="Email mismatch")
    if datetime.utcnow() > datetime.fromisoformat(session.get('otp_expiry')):
        return jsonify(success=False, message="OTP expired")
    if otp != session.get('admin_otp'):
        return jsonify(success=False, message="Invalid OTP")

    # Get user details to set the session correctly
    user = db.session.execute(db.select(User).where(func.lower(User.email) == email.lower())).scalar()

    if not user:
         return jsonify(success=False, message="User not found after OTP validation.")

    # Set all necessary session variables
    session['user'] = user.email
    session['role'] = user.role
    session['real_role'] = user.real_role # IMPORTANT: Store real_role for permissions

    session.pop('admin_otp', None)
    session.pop('otp_expiry', None)

    log_activity(email, "Admin Panel Login", "Logged into admin panel")

    return jsonify(success=True)


def send_email(to_email, subject, body):
    from_email = "shreyas22032004@gmail.com"
    password = "wvcnpazhztcutrmb"
    smtp_server = "smtp.gmail.com"
    port = 587

    message = f"Subject: {subject}\n\n{body}"

    with smtplib.SMTP(smtp_server, port) as server:
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, to_email, message)


@app.route('/check-email', methods=['POST'])
def check_email():
    data = request.get_json()
    email = data.get('email')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM users WHERE lower(email) = %s", (email.lower(),))
    exists = cur.fetchone() is not None
    cur.close()
    conn.close()

    return jsonify({'exists': exists})


@app.route('/admin', methods=['GET', 'POST'], endpoint='adminpage')
def admin():
    # 1. Check if the user is authorized to even see the admin page
    if 'user' not in session or session.get('real_role') not in ['Admin', 'Superadmin']:
        flash("You do not have permission to access this page.", "danger")
        return redirect(url_for('admin_login'))

    current_user_email = session.get('user')
    current_user_real_role = session.get('real_role')
    
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        real_role = request.form.get('real_role')
        role = request.form.get('role')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        # 2. Permission Check on User Creation: Admins cannot create other admins.
        if current_user_real_role == 'Admin' and real_role in ['Superadmin', 'Admin']:
            flash("Admins cannot create other Admin or Superadmin accounts.", "danger")
            cur.close()
            conn.close()
            return redirect('/admin')
        
        if password != confirm_password:
            flash("Passwords do not match", "danger")
            return redirect('/admin')

        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        registered_date = datetime.now()
        status = "Disabled" # New users always start as disabled

        try:
            cur.execute("""
                INSERT INTO users (real_role, role, email, password, registered_date, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (real_role, role, email, hashed_password, registered_date, status))
            conn.commit()
            
            log_activity(current_user_email, f"Add {real_role}", f"Name: {role}, Email: {email}")
            send_credentials_email(email, role, password)
            flash(f"User {role} created successfully. Credentials sent to {email}.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Database error: {e}", "danger")
        
        finally:
            cur.close()
            conn.close()
        
        return redirect('/admin')

    # --- GET Request Logic ---
    cur.execute("""
        SELECT real_role, role, email, registered_date, status
        FROM users
        ORDER BY registered_date DESC
    """)
    rows = cur.fetchall()
    users = [{"real_role": r[0], "role": r[1], "email": r[2], "registered_date": r[3], "status": r[4]} for r in rows]

    cur.execute("SELECT role FROM users WHERE lower(email) = %s", (current_user_email.lower(),))
    result = cur.fetchone()
    name = result[0] if result else "Admin"

    cur.close()
    conn.close()

    hour = datetime.now().hour
    greeting = ("Good Morning," if 5 <= hour < 12 else "Good Afternoon," if 12 <= hour < 17 else "Good Evening,")

    # 3. Pass the current user's role to the template for conditional rendering
    return render_template(
        'admin.html',
        users=users,
        current_email=current_user_email,
        greeting=greeting,
        name=name,
        current_user_real_role=current_user_real_role 
    )



@app.route('/edit-role', methods=['POST'])
def edit_role():
    if 'user' not in session:
        flash('Unauthorized', 'danger')
        return redirect('/login')

    email_to_edit = request.form.get('email')
    new_role = request.form.get('new_role')
    current_user_email = session.get('user')
    current_user_real_role = session.get('real_role')

    conn = get_db_connection()
    cur = conn.cursor()

    # Get the target user's current role to check against
    cur.execute("SELECT real_role FROM users WHERE email = %s", (email_to_edit,))
    target_user_role_tuple = cur.fetchone()
    
    if not target_user_role_tuple:
        flash("User to edit not found.", "danger")
        cur.close()
        conn.close()
        return redirect('/admin')
        
    target_user_role = target_user_role_tuple[0]
    
    # --- Permission Check using the helper function ---
    if not has_permission_to_edit(current_user_real_role, target_user_role):
        flash("🚫 You do not have permission to edit this user's role.", "danger")
        cur.close()
        conn.close()
        return redirect('/admin')

    # If permission is granted, update the role in the database
    cur.execute("UPDATE users SET real_role = %s WHERE email = %s", (new_role, email_to_edit))
    conn.commit()
    log_activity(current_user_email, "Edit Role", f"Changed {email_to_edit} from {target_user_role} to {new_role}")
    
    cur.close()
    conn.close()

    flash("✅ Role updated successfully", "success")
    return redirect('/admin')




@app.route('/delete-user', methods=['POST'])
def delete_user():
    if 'user' not in session:
        flash('Unauthorized', 'danger')
        return redirect('/login')

    email_to_delete = request.form.get('email')
    current_user_email = session.get('user')
    current_user_real_role = session.get('real_role')

    if email_to_delete == current_user_email:
        flash("You cannot delete yourself.", "danger")
        return redirect('/admin')

    conn = get_db_connection()
    cur = conn.cursor()

    # Get the target user's role to check permissions
    cur.execute("SELECT real_role FROM users WHERE email = %s", (email_to_delete,))
    target_user_role_tuple = cur.fetchone()
    
    if not target_user_role_tuple:
        flash("User to delete not found.", "danger")
        cur.close()
        conn.close()
        return redirect('/admin')
        
    target_user_role = target_user_role_tuple[0]
    
    # --- Permission Check using the helper function ---
    if not has_permission_to_edit(current_user_real_role, target_user_role):
        flash("🚫 You do not have permission to delete this user.", "danger")
        cur.close()
        conn.close()
        return redirect('/admin')

    # If permission is granted, proceed with deletion
    cur.execute("DELETE FROM users WHERE email = %s", (email_to_delete,))
    conn.commit()
    log_activity(current_user_email, "Delete User", f"Deleted user {email_to_delete} (Role: {target_user_role})")
    
    cur.close()
    conn.close()

    flash("✅ User deleted successfully", "success")
    return redirect('/admin')

@app.route("/api/toggle-user-status", methods=["POST"])
def toggle_user_status():
    if 'user' not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    data = request.get_json()
    email_to_toggle = data.get("email")
    new_status = data.get("status")

    current_user_email = session.get('user')
    current_user_real_role = session.get('real_role')

    # Add a backend check to prevent users from changing their own status
    if email_to_toggle == current_user_email:
        return jsonify({"success": False, "error": "You cannot change your own status."}), 403

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get the role of the user being toggled
        cur.execute("SELECT real_role FROM users WHERE email = %s", (email_to_toggle,))
        target_user_tuple = cur.fetchone()
        
        if not target_user_tuple:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Target user not found."}), 404
            
        target_user_role = target_user_tuple[0]

        # --- Permission Check using the existing helper function ---
        if not has_permission_to_edit(current_user_real_role, target_user_role):
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Permission denied."}), 403

        # If permission check passes, proceed with the update
        cur.execute("UPDATE users SET status = %s WHERE email = %s", (new_status, email_to_toggle))
        conn.commit()

        # Log the activity
        log_activity(
            current_user_email,
            f"{'Enable' if new_status == 'Enabled' else 'Disable'}",
            f"{target_user_role} Email: {email_to_toggle}"
        )

        cur.close()
        conn.close()
        return jsonify({"success": True})
        
    except Exception as e:
        # Log the exception e for debugging
        print(f"Error in toggle_user_status: {e}")
        return jsonify({"success": False, "error": "An internal error occurred."}), 500
    
import csv

@app.route('/download-users-csv')
def download_users_csv():
    selected_emails = request.args.getlist('emails')

    conn = get_db_connection()
    cur = conn.cursor()

    if selected_emails:
        # Generate placeholders for PostgreSQL
        placeholders = ','.join(['%s'] * len(selected_emails))
        query = f"""
            SELECT role, email, real_role, registered_date, status
            FROM users
            WHERE email IN ({placeholders})
            ORDER BY registered_date DESC
        """
        cur.execute(query, tuple(selected_emails))
    else:
        cur.execute("""
            SELECT role, email, real_role, registered_date, status
            FROM users
            ORDER BY registered_date DESC
        """)

    users = cur.fetchall()
    cur.close()
    conn.close()

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Email', 'Role', 'Registered Date', 'Status'])

    for row in users:
        writer.writerow([
            row[0],
            row[1],
            row[2],
            row[3].strftime('%d/%m/%Y') if row[3] else '',
            row[4]
        ])

    output.seek(0)
    log_activity(session.get('user'), "Download Users Table")
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='users.csv'
    )


@app.route("/api/admin-users")
def get_admin_users():
    current_email = session.get('user')
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Temporarily remove filtering
    cur.execute("""
        SELECT id, role, email, registered_date, status
        FROM users
        ORDER BY registered_date DESC
    """)
    
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify({
        "current_email": current_email,
        "admins": [
            {
                "id": row[0],
                "role": row[1],
                "email": row[2],
                "registered_date": row[3].strftime('%Y-%m-%d %H:%M:%S') if row[3] else None,
                "status": row[4]
            }
            for row in rows
        ]
    })

@app.route("/api/activity-log")
def api_activity_log():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            u.role,
            u.real_role,
            a.email,
            a.action,
            a.details,
            a.changes,
            a.timestamp
        FROM activity_log a
        LEFT JOIN users u ON LOWER(a.email) = LOWER(u.email)
        ORDER BY a.timestamp DESC
        LIMIT 200
    """)

    rows = cur.fetchall()
    cur.close()
    conn.close()

    logs = [{
        "role": r[0],
        "real_role": r[1],
        "email": r[2],
        "action": r[3],
        "details": r[4],
        "changes": r[5],
        "timestamp": r[6].strftime("%Y-%m-%dT%H:%M:%S")  # frontend will split date and time
    } for r in rows]

    return jsonify(logs)


@app.route('/activity-log')
def activity_log():
    return render_template('activity_log.html')

def log_activity(email, action, details=None, changes=None):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO activity_log (email, action, details, changes)
            VALUES (%s, %s, %s, %s)
        """, (email, action, details, changes))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print("❌ Failed to log activity:", e)
        
from user_agents import parse
@app.route("/api/ping", methods=["POST"])
def ping():
    data = request.get_json()
    email = session.get("user")
    page = data.get("page")

    if not email:
        return jsonify({"status": "unauthenticated"}), 401

    # Get device from user-agent
    user_agent_string = request.headers.get("User-Agent")
    ua = parse(user_agent_string)
    device = f"{ua.os.family} / {ua.browser.family}"  # Example: "Windows / Chrome"

    stmt = text("""
        INSERT INTO user_activity (email, page, last_seen, device)
        VALUES (:email, :page, NOW(), :device)
        ON CONFLICT (email)
        DO UPDATE SET page = EXCLUDED.page, last_seen = NOW(), device = EXCLUDED.device
    """)
    db.session.execute(stmt, {"email": email, "page": page, "device": device})
    db.session.commit()
    return jsonify({"status": "ok"})







@app.route("/api/live-users")
def live_users():
    result = db.session.execute(text("""
        SELECT u.role AS name, u.email, a.page, a.last_seen, a.device
        FROM users u
        LEFT JOIN user_activity a ON u.email = a.email
        ORDER BY u.email
    """)).fetchall()

    users = []
    for r in result:
        users.append({
            "name": r.name,
            "email": r.email,
            "page": r.page,
            "last_seen": r.last_seen.isoformat() if r.last_seen else None,
            "device": r.device or "Unknown"
        })

    return jsonify(users)

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_credentials_email(to_email, name, password):
    sender_email = "shreyas22032004@gmail.com"
    sender_password = "wvcnpazhztcutrmb"  # App password if using Gmail
    smtp_server = "smtp.gmail.com"
    smtp_port = 587

    subject = "Your Sales Order Tracker Account Details"

    html_body = f"""
    <html>
      <body>
        <p>Hi <strong>{name}</strong>,<br><br>
        Your account has been created on the Sales Order Tracker portal.<br><br>
        <b>Login Details:</b><br>
        Email: <code>{to_email}</code><br>
        Password: <code>{password}</code><br><br>
        You can log in at: <a href="http://yourdomain.com/login">http://yourdomain.com/login</a><br><br>
        Regards,<br>
        Admin Team
        </p>
      </body>
    </html>
    """

    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = to_email
    message.attach(MIMEText(html_body, "html"))

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, to_email, message.as_string())
        server.quit()
        print(f"✅ Email sent to {to_email}")
    except Exception as e:
        print(f"❌ Email sending failed: {e}")



@app.route('/api/kick-user', methods=['POST'])
def kick_user():
    data = request.get_json()
    kicked_email = data.get('email')
    admin_email = session.get('user')  # The admin performing the kick

    if not kicked_email:
        return jsonify({'success': False, 'message': 'Email is required'}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Mark user as kicked and disable their status
        cur.execute("""
            UPDATE users
            SET kicked = TRUE, status = 'Disabled'
            WHERE email = %s
        """, (kicked_email,))

        # Get kicked user's role for logging
        cur.execute("SELECT role FROM users WHERE email = %s", (kicked_email,))
        result = cur.fetchone()
        user_role = result[0] if result else "Unknown"

        # Log the kick in activity_log with admin's email
        cur.execute("""
            INSERT INTO activity_log (email, action, details, timestamp)
            VALUES (%s, %s, %s, NOW())
        """, (
            admin_email,
            'Kick',
            f"{user_role} ({kicked_email}) was forcibly logged out and disabled by admin"
        ))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({'success': True})

    except Exception as e:
        print("Kick user error:", e)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/check-session')
def check_session():
    user_email = session.get('user')
    if not user_email:
        return jsonify({'valid': False}), 401

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT kicked FROM users WHERE email = %s", (user_email,))
        kicked_status = cur.fetchone()
        cur.close()
        conn.close()

        if kicked_status and kicked_status[0]:
            return jsonify({'valid': False, 'message': 'User has been kicked out'}), 403

        return jsonify({'valid': True})
    except Exception as e:
        print("Check session error:", e)
        return jsonify({'valid': False, 'message': str(e)}), 500

@app.route('/api/send-expired-po-pdf', methods=['POST'])
def send_expired_po_pdf():
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    import json

    recipients = request.form.get('recipients')
    pdf_file = request.files.get('pdf')
    if not recipients or not pdf_file:
        return jsonify(success=False, error="Missing recipients or PDF file"), 400

    emails = json.loads(recipients)
    subject = "Expired Purchase Orders"
    body = "Please find attached the expired purchase orders report."

    sender_email = app.config.get('MAIL_USERNAME') or "shreyas22032004@gmail.com"
    sender_password = app.config.get('MAIL_PASSWORD') or "wvcnpazhztcutrmb"
    smtp_server = app.config.get('MAIL_SERVER') or "smtp.gmail.com"
    smtp_port = app.config.get('MAIL_PORT') or 587

    # Read PDF data once
    pdf_data = pdf_file.read()

    # If emails is empty, do nothing but return success
    if not emails:
        print("No recipients provided, skipping email send.")
        return jsonify(success=True)

    try:
        for to_email in emails:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = to_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain'))

            part = MIMEApplication(pdf_data, Name="ExpiredPOs.pdf")
            part['Content-Disposition'] = 'attachment; filename="ExpiredPOs.pdf"'
            msg.attach(part)

            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, to_email, msg.as_string())
                print(f"Sent expired PO PDF to {to_email}")

        return jsonify(success=True)
    except Exception as e:
        print("SMTP send error:", e)
        return jsonify(success=False, error=str(e)), 500

from flask import Flask, request, session, redirect
import jwt

@app.route('/portal_redirect')
def portal_redirect():
    user_email = session.get("user_email")
    role = session.get("role")

    if not user_email:
        return redirect("/login")

    # Create JWT token
    payload = {
        'email': user_email,
        'role': role,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(minutes=5)
    }
    token = jwt.encode(payload, JWT_SECRET_KEY, algorithm='HS256')

    return redirect(url_for('portal_auto_login', token=token))

# Portal receives token, auto-logs in
@app.route('/auto-login')
def portal_auto_login():
    token = request.args.get('token')
    if not token:
        return redirect('/portal-login')

    try:
        decoded = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
        session['user'] = decoded.get('email')
        session['role'] = decoded.get('role')
        return redirect('/dashboard')
    except jwt.ExpiredSignatureError:
        return "Token expired. Please log in manually.", 401
    except jwt.InvalidTokenError:
        return "Invalid token. Please log in.", 401

 ################ Po's Overview #############      

@app.route('/po-overview')
def po_overview():
    return render_template('po_overview_report.html')



@app.route('/api/po-performance-data', methods=['POST'])
def po_performance_data():
    
    from collections import defaultdict
    from datetime import datetime, date

    filters = request.get_json()

    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])
    interval = filters.get("interval", "Monthly")
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    group_by = filters.get("group_by", "period")

    def is_valid_date(date_str):
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except:
            return False

    valid_from = is_valid_date(from_date)
    valid_to = is_valid_date(to_date)

    where_clauses = []
    params = []

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            # ✅ Determine default from/to dates
            if not valid_from or not valid_to:
                cur.execute("SELECT MIN(po_date) FROM orders")
                min_po_date = cur.fetchone()[0]
                default_from = min_po_date if min_po_date else date.today()
                default_to = date.today()

                if not valid_from:
                    from_date = default_from.strftime("%Y-%m-%d")
                if not valid_to:
                    to_date = default_to.strftime("%Y-%m-%d")
            else:
                default_from = datetime.strptime(from_date, "%Y-%m-%d").date()
                default_to = datetime.strptime(to_date, "%Y-%m-%d").date()

            # Apply filters
            if company:
                where_clauses.append("company_name = ANY(%s)")
                params.append(company)
            if client:
                where_clauses.append("client_name = ANY(%s)")
                params.append(client)
            if division:
                where_clauses.append("division = ANY(%s)")
                params.append(division)
            if subcontractor:
                where_clauses.append("sub_contractor_name = ANY(%s)")
                params.append(subcontractor)
            if site:
                where_clauses.append("site = ANY(%s)")
                params.append(site)

            where_clauses.append("po_date BETWEEN %s AND %s")
            params.extend([from_date, to_date])

            query = """
                SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                       start_date, end_date, site, po_number, po_amount, issued_by,
                       po_date, status, project_duration, po_expiry_date
                FROM orders
                WHERE TRUE
            """
            if where_clauses:
                query += " AND " + " AND ".join(where_clauses)

            cur.execute(query, params)
            rows = cur.fetchall()
            cols = [desc[0] for desc in cur.description]
            results = []
            for row in rows:
                record = dict(zip(cols, row))
                record["order_id"] = record.pop("id", None)
                results.append(record)

    entity_groupings = [
        "company_name", "client_name", "division", "site", "sub_contractor_name"
    ]

    if group_by in entity_groupings:
        # For entity-based grouping
        output = []
        for row in results:
            full_entity = row.get(group_by) or "Unknown"
            short_entity = " ".join([word[0].upper() for word in full_entity.split() if word and word[0].isalpha()])
            date_obj = row.get("po_date")

            if date_obj:
                if interval == "Weekly":
                    time_label = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
                elif interval == "Quarterly":
                    time_label = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
                elif interval == "Yearly":
                    time_label = f"{date_obj.year}"
                else:
                    time_label = date_obj.strftime("%b %Y")
                combined_label = f"{short_entity} - {time_label}"
            else:
                combined_label = short_entity

            output.append({
                "period": combined_label,
                "po": row["po_number"],
                "y": float(row["po_amount"] or 0),
                "project": row.get("project_name") or "-",
                "entity_name": full_entity,
                "status": row.get("status", "Unknown"),
                "po_date": date_obj.strftime("%Y-%m-%d") if date_obj else None
            })

        def extract_date(label):
            try:
                timeline = label.split(" - ")[-1]
                if interval == "Weekly":
                    year, week = timeline.split(" W")
                    return datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w")
                elif interval == "Quarterly":
                    q, y = timeline.split()
                    month = (int(q[1]) - 1) * 3 + 1
                    return datetime(int(y), month, 1)
                elif interval == "Yearly":
                    return datetime(int(timeline), 1, 1)
                else:
                    return datetime.strptime(timeline, "%b %Y")
            except:
                return datetime.max

        output.sort(key=lambda x: extract_date(x["period"]))

        return jsonify({
            "data": output,
            "records": results,
            "default_from_date": default_from.strftime("%Y-%m-%d"),
            "default_to_date": default_to.strftime("%Y-%m-%d")
        })

    # Timeline grouping with status
    summary = defaultdict(lambda: {"total_sales": 0, "sales_count": 0, "status_counts": defaultdict(int)})
    for row in results:
        date_obj = row.get("po_date")
        if not date_obj:
            continue
        if interval == "Weekly":
            key = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
        elif interval == "Quarterly":
            key = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
        elif interval == "Yearly":
            key = f"{date_obj.year}"
        else:
            key = date_obj.strftime("%b %Y")

        status = (row.get("status") or "Unknown").lower()
        summary[key]["total_sales"] += float(row["po_amount"] or 0)
        summary[key]["sales_count"] += 1
        summary[key]["status_counts"][status] += 1

    output = []
    for period, data in sorted(summary.items()):
        output.append({
            "period": period,
            "total_sales": data["total_sales"],
            "sales_count": data["sales_count"],
            "statuses": dict(data["status_counts"])
        })

    return jsonify({
        "data": output,
        "records": results,
        "default_from_date": default_from.strftime("%Y-%m-%d"),
        "default_to_date": default_to.strftime("%Y-%m-%d")
    })




@app.route('/api/filter-options') 
def po_filter_options():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    # Build WHERE clause for timeline
    where_clauses = []
    params = []

    if from_date and to_date:
        where_clauses.append("po_date BETWEEN %s AND %s")
        params.extend([from_date, to_date])

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT DISTINCT company_name FROM orders {where_sql} ORDER BY company_name ASC;", params)
            company_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT client_name FROM orders {where_sql} ORDER BY client_name ASC;", params)
            client_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT sub_contractor_name FROM orders WHERE sub_contractor_name IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY sub_contractor_name ASC",
                        params if from_date and to_date else [])
            subcontractors = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT site FROM orders WHERE site IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY site ASC",
                        params if from_date and to_date else [])
            sites = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT division FROM orders {where_sql} ORDER BY division ASC;", params)
            divisions = [row[0] for row in cur.fetchall()]

    return jsonify({
        "company_names": company_names,
        "client_names": client_names,
        "subcontractors": subcontractors,
        "sites": sites,
        "divisions": divisions
    })


@app.route("/api/filter-options-dependent-po", methods=["POST"])
def filter_options_dependent_po():
    filters = request.json or {}
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])

    where = []
    params = []

    if company:
        where.append("company_name = ANY(%s)")
        params.append(company)
    if client:
        where.append("client_name = ANY(%s)")
        params.append(client)
    if division:
        where.append("division = ANY(%s)")
        params.append(division)
    if subcontractor:
        where.append("sub_contractor_name = ANY(%s)")
        params.append(subcontractor)
    if site:
        where.append("site = ANY(%s)")
        params.append(site)
    if from_date:
        where.append("po_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("po_date <= %s")
        params.append(to_date)

    query = """
        SELECT DISTINCT client_name, division, sub_contractor_name, site, company_name
        FROM orders
    """
    if where:
        query += " WHERE " + " AND ".join(where)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Extract unique values for each field
    clients = sorted(set(r[0] for r in rows if r[0]))
    divisions = sorted(set(r[1] for r in rows if r[1]))
    subcontractors = sorted(set(r[2] for r in rows if r[2]))
    sites = sorted(set(r[3] for r in rows if r[3]))
    companies = sorted(set(r[4] for r in rows if r[4]))

    return jsonify({
        "client_names": clients,
        "divisions": divisions,
        "subcontractors": subcontractors,
        "sites": sites,
        "company_names": companies
    })



@app.route("/api/get-po-date-range-alt")
def get_po_date_range_alt():
    from datetime import date
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT MIN(po_date) FROM orders")
        min_date = cur.fetchone()[0]
        cur.close()
        conn.close()

        return jsonify({
            "min_date": min_date.strftime('%Y-%m-%d') if min_date else date.today().strftime('%Y-%m-%d'),
            "max_date": date.today().strftime('%Y-%m-%d')  # ✅ always today
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500    



# API: PO Overview Data
@app.route('/api/po-overview')
def po_overview_data():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT po_number, SUM(po_amount) as total_sales, COUNT(*) as order_count
        FROM orders
        GROUP BY po_number
        ORDER BY total_sales DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    data = [{
        "po_number": row[0],
        "total_sales": row[1],
        "order_count": row[2]
    } for row in rows]

    return jsonify(data)   


@app.route('/api/po-filters')
def get_po_filters():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT company_name FROM orders")
    companies = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT client_name FROM orders")
    clients = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT sub_contractor_name FROM orders")
    subcontractors = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT site FROM orders")
    sites = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT DISTINCT division FROM orders")
    divisions = [row[0] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return jsonify({
        'companies': companies,
        'clients': clients,
        'subcontractors': subcontractors,
        'sites': sites,
        'divisions': divisions
    })



@app.route('/api/po-overview-data', methods=['POST'])
def po_overview_data_filtered():
    from datetime import datetime # Make sure datetime is imported

    data = request.get_json()
    if not data:
        return jsonify({"error": "Invalid request"}), 400

    # ✅ THE FIX: The incorrect date parsing function has been removed.
    # We now get the date strings directly from the frontend request.
    # The database can handle the 'YYYY-MM-DD' format automatically.
    
    companies = data.get('company_name', [])
    clients = data.get('client_name', [])
    subcontractors = data.get('sub_contractor_name', [])
    sites = data.get('site', [])
    divisions = data.get('division', [])
    from_date = data.get('from_date')
    to_date = data.get('to_date')
    sort_key = data.get('sort', '')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    base_query = "FROM orders WHERE 1=1"
    params = []

    if companies:
        base_query += " AND company_name = ANY(%s)"
        params.append(companies)
    if clients:
        base_query += " AND client_name = ANY(%s)"
        params.append(clients)
    if subcontractors:
        base_query += " AND sub_contractor_name = ANY(%s)"
        params.append(subcontractors)
    if sites:
        base_query += " AND site = ANY(%s)"
        params.append(sites)
    if divisions:
        base_query += " AND division = ANY(%s)"
        params.append(divisions)
        
    # This logic correctly handles the dates now
    if from_date:
        base_query += " AND po_date >= %s"
        params.append(from_date)
    if to_date:
        base_query += " AND po_date <= %s"
        params.append(to_date)

    # --- Graph Data Query ---
    graph_query = f"""
        SELECT po_number, SUM(po_amount) AS total_sales, COUNT(*) AS order_count, MIN(po_date) as po_date
        {base_query}
        GROUP BY po_number
    """
    if sort_key == 'sales_desc':
        graph_query += " ORDER BY total_sales DESC"
    elif sort_key == 'sales_asc':
        graph_query += " ORDER BY total_sales ASC"
    else:
        graph_query += " ORDER BY po_number"

    cur.execute(graph_query, tuple(params))
    graph_rows = cur.fetchall()
    
    graph_data = []
    for row in graph_rows:
        graph_data.append({
            'po_number': row['po_number'],
            'total_sales': row['total_sales'],
            'order_count': row['order_count'],
            'po_date': row['po_date'].strftime('%d-%b-%Y') if row['po_date'] else 'N/A'
        })
        
    # --- Details Table Query ---
    table_query = f"SELECT po_number, po_date, po_expiry_date, client_name, company_name, project_name, site, division, po_amount, status, issued_by {base_query}"
    cur.execute(table_query, tuple(params))
    records = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify({
        'graph_data': graph_data,
        'records': records
    })


 ################ ROHAN'S NEW SAMRAJYA #############      

@app.route('/growth-overview')
def growth_overview():
    return render_template('growth_overview.html')



@app.route('/api/growth-performance-data', methods=['POST'])
def growth_performance_data():
    
    from collections import defaultdict
    from datetime import datetime, date

    filters = request.get_json()

    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])
    interval = filters.get("interval", "Monthly")
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    group_by = filters.get("group_by", "period")

    def is_valid_date(date_str):
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except:
            return False

    valid_from = is_valid_date(from_date)
    valid_to = is_valid_date(to_date)

    where_clauses = []
    params = []

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            # ✅ Determine default from/to dates
            if not valid_from or not valid_to:
                cur.execute("SELECT MIN(po_date) FROM orders")
                min_po_date = cur.fetchone()[0]
                default_from = min_po_date if min_po_date else date.today()
                default_to = date.today()

                if not valid_from:
                    from_date = default_from.strftime("%Y-%m-%d")
                if not valid_to:
                    to_date = default_to.strftime("%Y-%m-%d")
            else:
                default_from = datetime.strptime(from_date, "%Y-%m-%d").date()
                default_to = datetime.strptime(to_date, "%Y-%m-%d").date()

            # Apply filters
            if company:
                where_clauses.append("company_name = ANY(%s)")
                params.append(company)
            if client:
                where_clauses.append("client_name = ANY(%s)")
                params.append(client)
            if division:
                where_clauses.append("division = ANY(%s)")
                params.append(division)
            if subcontractor:
                where_clauses.append("sub_contractor_name = ANY(%s)")
                params.append(subcontractor)
            if site:
                where_clauses.append("site = ANY(%s)")
                params.append(site)

            where_clauses.append("po_date BETWEEN %s AND %s")
            params.extend([from_date, to_date])

            query = """
                SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                       start_date, end_date, site, po_number, po_amount, issued_by,
                       po_date, status, project_duration, po_expiry_date
                FROM orders
                WHERE TRUE
            """
            if where_clauses:
                query += " AND " + " AND ".join(where_clauses)

            cur.execute(query, params)
            rows = cur.fetchall()
            cols = [desc[0] for desc in cur.description]
            results = []
            for row in rows:
                record = dict(zip(cols, row))
                record["order_id"] = record.pop("id", None)
                results.append(record)

    entity_groupings = [
        "company_name", "client_name", "division", "site", "sub_contractor_name"
    ]

    if group_by in entity_groupings:
        # For entity-based grouping
        output = []
        for row in results:
            full_entity = row.get(group_by) or "Unknown"
            short_entity = " ".join([word[0].upper() for word in full_entity.split() if word and word[0].isalpha()])
            date_obj = row.get("po_date")

            if date_obj:
                if interval == "Weekly":
                    time_label = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
                elif interval == "Quarterly":
                    time_label = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
                elif interval == "Yearly":
                    time_label = f"{date_obj.year}"
                else:
                    time_label = date_obj.strftime("%b %Y")
                combined_label = f"{short_entity} - {time_label}"
            else:
                combined_label = short_entity

            output.append({
                "period": combined_label,
                "po": row["po_number"],
                "y": float(row["po_amount"] or 0),
                "project": row.get("project_name") or "-",
                "entity_name": full_entity,
                "status": row.get("status", "Unknown"),
                "po_date": date_obj.strftime("%Y-%m-%d") if date_obj else None
            })

        def extract_date(label):
            try:
                timeline = label.split(" - ")[-1]
                if interval == "Weekly":
                    year, week = timeline.split(" W")
                    return datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w")
                elif interval == "Quarterly":
                    q, y = timeline.split()
                    month = (int(q[1]) - 1) * 3 + 1
                    return datetime(int(y), month, 1)
                elif interval == "Yearly":
                    return datetime(int(timeline), 1, 1)
                else:
                    return datetime.strptime(timeline, "%b %Y")
            except:
                return datetime.max

        output.sort(key=lambda x: extract_date(x["period"]))

        return jsonify({
            "data": output,
            "records": results,
            "default_from_date": default_from.strftime("%Y-%m-%d"),
            "default_to_date": default_to.strftime("%Y-%m-%d")
        })

    # Timeline grouping with status
    summary = defaultdict(lambda: {"total_sales": 0, "sales_count": 0, "status_counts": defaultdict(int)})
    for row in results:
        date_obj = row.get("po_date")
        if not date_obj:
            continue
        if interval == "Weekly":
            key = f"{date_obj.year} W{date_obj.isocalendar()[1]}"
        elif interval == "Quarterly":
            key = f"Q{(date_obj.month - 1) // 3 + 1} {date_obj.year}"
        elif interval == "Yearly":
            key = f"{date_obj.year}"
        else:
            key = date_obj.strftime("%b %Y")

        status = (row.get("status") or "Unknown").lower()
        summary[key]["total_sales"] += float(row["po_amount"] or 0)
        summary[key]["sales_count"] += 1
        summary[key]["status_counts"][status] += 1

    output = []
    for period, data in sorted(summary.items()):
        output.append({
            "period": period,
            "total_sales": data["total_sales"],
            "sales_count": data["sales_count"],
            "statuses": dict(data["status_counts"])
        })

    return jsonify({
        "data": output,
        "records": results,
        "default_from_date": default_from.strftime("%Y-%m-%d"),
        "default_to_date": default_to.strftime("%Y-%m-%d")
    })



@app.route('/api/growth-filter-options')
def growth_filter_options():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    # Build WHERE clause for timeline
    where_clauses = []
    params = []

    if from_date and to_date:
        where_clauses.append("po_date BETWEEN %s AND %s")
        params.extend([from_date, to_date])

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT DISTINCT company_name FROM orders {where_sql} ORDER BY company_name ASC;", params)
            company_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT client_name FROM orders {where_sql} ORDER BY client_name ASC;", params)
            client_names = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT sub_contractor_name FROM orders WHERE sub_contractor_name IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY sub_contractor_name ASC",
                        params if from_date and to_date else [])
            subcontractors = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT site FROM orders WHERE site IS NOT NULL" +
                        (f" AND po_date BETWEEN %s AND %s" if from_date and to_date else "") +
                        " ORDER BY site ASC",
                        params if from_date and to_date else [])
            sites = [row[0] for row in cur.fetchall()]

            cur.execute(f"SELECT DISTINCT division FROM orders {where_sql} ORDER BY division ASC;", params)
            divisions = [row[0] for row in cur.fetchall()]

    return jsonify({
        "company_names": company_names,
        "client_names": client_names,
        "subcontractors": subcontractors,
        "sites": sites,
        "divisions": divisions
    })



@app.route("/api/growth-filter-options-dependent", methods=["POST"])
def growth_filter_options_dependent():
    filters = request.json or {}
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")
    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])

    where = []
    params = []

    if company:
        where.append("company_name = ANY(%s)")
        params.append(company)
    if client:
        where.append("client_name = ANY(%s)")
        params.append(client)
    if division:
        where.append("division = ANY(%s)")
        params.append(division)
    if subcontractor:
        where.append("sub_contractor_name = ANY(%s)")
        params.append(subcontractor)
    if site:
        where.append("site = ANY(%s)")
        params.append(site)
    if from_date:
        where.append("po_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("po_date <= %s")
        params.append(to_date)

    query = """
        SELECT DISTINCT client_name, division, sub_contractor_name, site, company_name
        FROM orders
    """
    if where:
        query += " WHERE " + " AND ".join(where)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Extract unique values for each field
    clients = sorted(set(r[0] for r in rows if r[0]))
    divisions = sorted(set(r[1] for r in rows if r[1]))
    subcontractors = sorted(set(r[2] for r in rows if r[2]))
    sites = sorted(set(r[3] for r in rows if r[3]))
    companies = sorted(set(r[4] for r in rows if r[4]))

    return jsonify({
        "client_names": clients,
        "divisions": divisions,
        "subcontractors": subcontractors,
        "sites": sites,
        "company_names": companies
    })


@app.route("/api/growth-order-details", methods=["POST"])
def get_growth_order_details():
    from datetime import datetime, timedelta
    data = request.get_json()
    group_by = data.get("group_by")
    value = data.get("value")
    interval = data.get("interval", "Monthly")

    if not group_by or value is None:
        return jsonify([])

    conn = get_db_connection()
    cur = conn.cursor()

    if group_by == "po_date":
        try:
            value_date = datetime.strptime(value, "%Y-%m-%d")
        except Exception:
            cur.close()
            conn.close()
            return jsonify([])

        from dateutil.relativedelta import relativedelta

        if interval == "Monthly":
            start = value_date.replace(day=1)
            end = start + relativedelta(months=1)
                

        elif interval == "Quarterly":
            quarter = (value_date.month - 1) // 3 + 1
            start_month = 3 * (quarter - 1) + 1
            start = value_date.replace(month=start_month, day=1)
            if start_month + 3 > 12:
                end = start.replace(year=start.year + 1, month=1, day=1)
            else:
                end = start.replace(month=start_month + 3, day=1)

        elif interval == "Yearly":
            start = value_date.replace(month=1, day=1)
            end = start.replace(year=start.year + 1)

        elif interval == "Weekly":
            start = value_date  # ISO week Monday already
            end = start + timedelta(days=7)

        else:
            cur.close()
            conn.close()
            return jsonify([])

        query = """
            SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                   start_date, end_date, site, po_number, po_amount, issued_by,
                   po_date, status, project_duration, po_expiry_date
            FROM orders
            WHERE DATE(po_date) >= %s AND DATE(po_date) < %s
        """
        cur.execute(query, (start.date(), end.date()))

    else:
        # For entity groupings (e.g., company_name, client_name, etc.)
        query = f"""
            SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
                   start_date, end_date, site, po_number, po_amount, issued_by,
                   po_date, status, project_duration, po_expiry_date
            FROM orders
            WHERE {group_by} = %s
        """
        cur.execute(query, (value,))

    rows = cur.fetchall()
    cols = [desc[0] for desc in cur.description]
    result = [dict(zip(cols, row)) for row in rows]

    cur.close()
    conn.close()
    return jsonify(result)

@app.route("/api/po-filter-options-dependent", methods=["POST"])
def po_filter_options_dependent():
    """
    This endpoint receives the current state of all filters and returns
    the relevant, filtered options for every dropdown.
    """
    filters = request.json or {}
    
    # Get the currently selected values from the frontend
    company = filters.get("company_name", [])
    client = filters.get("client_name", [])
    division = filters.get("division", [])
    subcontractor = filters.get("sub_contractor_name", [])
    site = filters.get("site", [])
    from_date = filters.get("from_date")
    to_date = filters.get("to_date")

    where = []
    params = []

    # Build the WHERE clause based on which filters are active
    if company:
        where.append("company_name = ANY(%s)")
        params.append(company)
    if client:
        where.append("client_name = ANY(%s)")
        params.append(client)
    if division:
        where.append("division = ANY(%s)")
        params.append(division)
    if subcontractor:
        where.append("sub_contractor_name = ANY(%s)")
        params.append(subcontractor)
    if site:
        where.append("site = ANY(%s)")
        params.append(site)
    if from_date:
        where.append("po_date >= %s")
        params.append(from_date)
    if to_date:
        where.append("po_date <= %s")
        params.append(to_date)

    # Base query to find all possible combinations based on the filters
    query = "SELECT DISTINCT client_name, division, sub_contractor_name, site, company_name FROM orders"
    if where:
        query += " WHERE " + " AND ".join(where)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Extract the unique values for each category from the filtered results
    clients_set = sorted(set(r[0] for r in rows if r[0]))
    divisions_set = sorted(set(r[1] for r in rows if r[1]))
    subcontractors_set = sorted(set(r[2] for r in rows if r[2]))
    sites_set = sorted(set(r[3] for r in rows if r[3]))
    companies_set = sorted(set(r[4] for r in rows if r[4]))

    # Return the filtered lists as JSON
    return jsonify({
        "client_names": clients_set,
        "divisions": divisions_set,
        "subcontractors": subcontractors_set,
        "sites": sites_set,
        "company_names": companies_set
    })

@app.route("/api/po-report-date-range")
def get_po_report_date_range():
    """
    Calculates and returns the default date range for the PO report.
    'min_date' is the oldest PO date in the database.
    'max_date' is always today's date.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Query to find the earliest PO date from all orders
        cur.execute("SELECT MIN(po_date) FROM orders WHERE po_date IS NOT NULL")
        min_date_result = cur.fetchone()
        min_date = min_date_result[0] if min_date_result else None
        
        cur.close()
        conn.close()

        # Format dates to YYYY-MM-DD string, defaulting to today if no orders exist
        min_date_str = min_date.strftime('%Y-%m-%d') if min_date else date.today().strftime('%Y-%m-%d')
        max_date_str = date.today().strftime('%Y-%m-%d')

        return jsonify({
            "min_date": min_date_str,
            "max_date": max_date_str
        })
    except Exception as e:
        print(f"Error in /api/po-report-date-range: {e}")
        return jsonify({"error": str(e)}), 500


 ################ END OF ROHAN'S NEW SAMRAJYA #############     

 ###################SITE PERFORMANCE######################
@app.before_request
def before_request():
    g.start = time()
    site_metrics["total_requests"] += 1
    site_metrics["endpoint_hits"][request.path] += 1
    user = session.get("user")
    if user:
        site_metrics["connected_users"].add(user)

@app.after_request
def after_request(response):
    duration = time() - g.start
    site_metrics["latencies"].append(duration)
    if len(site_metrics["latencies"]) > 100:
        site_metrics["latencies"] = site_metrics["latencies"][-100:]
    return response

@app.teardown_request
def on_teardown(exception):
    if exception:
        site_metrics["errors"] += 1
@app.route("/api/live-performance")
def get_live_performance():
    def get_avg(name):
        values = web_vital_metrics.get(name, [])
        return round(sum(values)/len(values), 2) if values else None

    avg_latency = sum(site_metrics["latencies"]) / len(site_metrics["latencies"]) if site_metrics["latencies"] else 0
    error_rate = (site_metrics["errors"] / site_metrics["total_requests"]) * 100 if site_metrics["total_requests"] else 0

    return jsonify({
        "Latency (ms)": round(avg_latency * 1000, 2),
        "Total Requests": site_metrics["total_requests"],
        "Error Rate (%)": round(error_rate, 2),
        "Live Users": len(site_metrics["connected_users"]),
        "Unique Endpoints": len(site_metrics["endpoint_hits"]),
        "Top Endpoint": max(site_metrics["endpoint_hits"], key=site_metrics["endpoint_hits"].get, default="-"),
        "FCP": get_avg("FCP"),
        "LCP": get_avg("LCP"),
        "TTFB": get_avg("TTFB"),
        "CLS": get_avg("CLS"),
        "INP": get_avg("INP"),
    })
@app.route('/api/track-web-vitals', methods=['POST'])
def track_web_vitals():
    data = request.get_json()
    name = data.get('name')
    value = float(data.get('value', 0))

    if name:
        web_vital_metrics[name].append(value)
        if len(web_vital_metrics[name]) > 100:
            web_vital_metrics[name] = web_vital_metrics[name][-100:]
    return '', 204


@app.route('/api/get-all-user-emails')
def get_all_user_emails():
    """
    Fetches a list of all user emails from the database.
    """
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # Query to get all emails from the users table
        cur.execute("SELECT email FROM users ORDER BY email ASC")
        users = cur.fetchall()
        cur.close()
        conn.close()

        # Extract emails from the query result
        email_list = [user[0] for user in users]
        
        return jsonify(email_list)
    except Exception as e:
        print(f"Error in /api/get-all-user-emails: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/generate-support-ticket')
def generate_support_ticket():
    """
    Generates a new, unique, sequential support ticket ID.
    It reads the last ID from a file, increments it, and saves it back.
    """
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    # This is the file where we will store the last ticket number
    ticket_counter_file = 'support_ticket_counter.txt'
    
    # --- Read the last ticket number ---
    try:
        with open(ticket_counter_file, 'r') as f:
            last_id = int(f.read().strip())
    except (FileNotFoundError, ValueError):
        # If the file doesn't exist or is empty, start from 0
        last_id = 0
        
    # --- Increment the number and save it back to the file ---
    new_id = last_id + 1
    with open(ticket_counter_file, 'w') as f:
        f.write(str(new_id))
        
    # --- Format the ticket ID and return it ---
    ticket_id = f"SOTP-{new_id:03d}" # e.g., SOTP-001, SOTP-012, SOTP-123
    
    return jsonify({"ticket_id": ticket_id})




if __name__ == "__main__":
    print("🚀 Starting Flask server...")
    app.run(host="0.0.0.0", port=5000, debug=True)


    
   
