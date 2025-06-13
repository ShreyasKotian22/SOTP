from flask import Flask, render_template, jsonify
from flask import send_file
import psycopg2
import tempfile
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

app = Flask(__name__)

# PostgreSQL configs
DB_CONFIG = {
    'dbname': 'login_db',
    'user': 'postgres',
    'password': 'shreyas',
    'host': 'localhost',
    'port': '5432'
}

def get_orders():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    # Only selecting needed columns
    cur.execute("""
        SELECT id, client_name, company_name, project_name, po_amount
        FROM orders
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Add dummy profit (you can later calculate this)
    result = []
    for row in rows:
        result.append({
            'order_id': row[0],
            'client_name': row[1],
            'company': row[2],
            'project': row[3],
            'po_amount': row[4],
            'profit': "₹0.00"  # Placeholder
        })
    return result

@app.route('/api/delete_order/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    try:
        conn = psycopg2.connect(
            dbname="login_db",
            user="postgres",
            password="shreyas",
            host="localhost",
            port="5432"
        )
        cur = conn.cursor()
        cur.execute("DELETE FROM orders WHERE id = %s", (order_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Order deleted"}), 200
    except Exception as e:
        print("Error deleting order:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/download/order/<int:order_id>')
def download_order_excel(order_id):
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, client_name, division, company_name, project_name, sub_contractor_name,
               start_date, end_date, site, po_number, po_amount, issued_by, po_date
        FROM orders
        WHERE id = %s
    """, (order_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return "Order not found", 404

    columns = [
        "Order ID", "Client Name", "Division", "Company Name", "Project Name", "Sub-Contractor Name",
        "Start Date", "End Date", "Site", "PO Number", "PO Amount", "Issued By", "PO Date"
    ]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Order {order_id}"

    # Write headers (row 1)
    for col_index, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)  # Make header bold

    # Write values (row 2)
    for col_index, value in enumerate(row, start=1):
        ws.cell(row=2, column=col_index, value=value)

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)

    return send_file(temp_file.name, as_attachment=True, download_name=f'Order_{order_id}.xlsx')    

@app.route('/')
def index():
    return render_template('view_sales_orders.html')

@app.route('/api/orders')
def api_orders():
    return jsonify(get_orders())

if __name__ == '__main__':
    app.run(debug=True)